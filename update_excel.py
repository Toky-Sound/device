import json
import openpyxl

# 讀取 JSON 資料
with open('data.json', 'r') as file:
    data = json.load(file)

# 讀取或創建 Excel 文件
file_path = 'health_data.xlsx'
try:
    workbook = openpyxl.load_workbook(file_path)
except FileNotFoundError:
    workbook = openpyxl.Workbook()

# 獲取工作表，或創建新的
if 'Sheet1' not in workbook.sheetnames:
    sheet = workbook.create_sheet('Sheet1')
    sheet.append(['血壓 (mmHg)', '血氧 (%)', '體脂肪 (%)', '時間'])
else:
    sheet = workbook['Sheet1']

# 添加新資料
sheet.append([data['bloodPressure'], data['bloodOxygen'], data['bodyFat'], data['timestamp']])

# 保存文件
workbook.save(file_path)
